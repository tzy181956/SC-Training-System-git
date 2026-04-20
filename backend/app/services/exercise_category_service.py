from __future__ import annotations

from pathlib import Path
from re import sub

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.exceptions import not_found
from app.models import Exercise, ExerciseCategory, ExerciseTag, Tag
from app.schemas.exercise_category import ExerciseImportPreview


EXOS_SOURCE_PATH = Path(r"C:\Users\tzy\Downloads\exos_action_library_zh_localized.xlsx")


def list_categories(db: Session) -> list[ExerciseCategory]:
    return db.query(ExerciseCategory).order_by(ExerciseCategory.level, ExerciseCategory.sort_order, ExerciseCategory.name_zh).all()


def list_category_tree(db: Session) -> list[ExerciseCategory]:
    categories = list_categories(db)
    category_map = {category.id: category for category in categories}
    for category in categories:
        category.children = []
    roots: list[ExerciseCategory] = []
    for category in categories:
        if category.parent_id and category.parent_id in category_map:
            category_map[category.parent_id].children.append(category)
        else:
            roots.append(category)
    return roots


def _slugify(value: str) -> str:
    value = sub(r"[^a-zA-Z0-9\u4e00-\u9fff]+", "-", value.strip().lower())
    value = sub(r"-{2,}", "-", value).strip("-")
    return value or "item"


def _build_code(parent_code: str | None, name_zh: str, name_en: str | None) -> str:
    local = _slugify(name_en or name_zh)
    return f"{parent_code}/{local}" if parent_code else local


def _iter_excel_rows(source_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook["中英对照"]
    headers = [str(cell or "").strip() for cell in next(sheet.iter_rows(values_only=True))]
    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        payload = {
            headers[index]: str(value).strip() if value is not None else ""
            for index, value in enumerate(row)
            if index < len(headers)
        }
        if payload.get("动作名称_中文"):
            rows.append(payload)
    return rows


def preview_exos_import(source_path: Path | None = None) -> ExerciseImportPreview:
    path = source_path or EXOS_SOURCE_PATH
    rows = _iter_excel_rows(path)

    level1 = set()
    level2 = set()
    level3 = set()
    exercises = set()
    skipped_duplicates = 0

    for row in rows:
        key1 = (row["一级分类_中文"], row["MovementTypeCategory_EN"])
        key2 = (*key1, row["二级分类_中文"], row["MovementType_EN"])
        key3 = (*key2, row["基础动作_中文"], row["BaseMovement_EN"])
        exercise_key = (*key3, row["动作名称_中文"], row["Movement_EN"])
        level1.add(key1)
        level2.add(key2)
        level3.add(key3)
        if exercise_key in exercises:
            skipped_duplicates += 1
        exercises.add(exercise_key)

    return ExerciseImportPreview(
        source_path=str(path),
        level1_categories=len(level1),
        level2_categories=len(level2),
        level3_categories=len(level3),
        exercises=len(exercises),
        skipped_duplicates=skipped_duplicates,
    )


def _get_or_create_category(
    db: Session,
    cache: dict[tuple[int | None, int, str], ExerciseCategory],
    *,
    parent: ExerciseCategory | None,
    level: int,
    name_zh: str,
    name_en: str | None,
    sort_order: int,
) -> ExerciseCategory:
    cache_key = (parent.id if parent else None, level, name_zh)
    if cache_key in cache:
        return cache[cache_key]

    category = (
        db.query(ExerciseCategory)
        .filter(
            ExerciseCategory.parent_id == (parent.id if parent else None),
            ExerciseCategory.level == level,
            ExerciseCategory.name_zh == name_zh,
        )
        .first()
    )
    if not category:
        category = ExerciseCategory(
            parent_id=parent.id if parent else None,
            level=level,
            name_zh=name_zh,
            name_en=name_en or None,
            code=_build_code(parent.code if parent else None, name_zh, name_en),
            sort_order=sort_order,
            is_system=True,
        )
        db.add(category)
        db.flush()
    cache[cache_key] = category
    return category


def import_exos_library(db: Session, *, source_path: Path | None = None, replace_existing: bool = True) -> ExerciseImportPreview:
    path = source_path or EXOS_SOURCE_PATH
    preview = preview_exos_import(path)
    rows = _iter_excel_rows(path)

    if replace_existing:
        db.query(ExerciseTag).delete()
        db.query(Exercise).delete()
        db.query(Tag).delete()
        db.query(ExerciseCategory).delete()
        db.flush()

    category_cache: dict[tuple[int | None, int, str], ExerciseCategory] = {}
    exercise_keys: set[tuple[int, str, str]] = set()

    for index, row in enumerate(rows, start=1):
        level1 = _get_or_create_category(
            db,
            category_cache,
            parent=None,
            level=1,
            name_zh=row["一级分类_中文"],
            name_en=row["MovementTypeCategory_EN"],
            sort_order=index,
        )
        level2 = _get_or_create_category(
            db,
            category_cache,
            parent=level1,
            level=2,
            name_zh=row["二级分类_中文"],
            name_en=row["MovementType_EN"],
            sort_order=index,
        )
        level3 = _get_or_create_category(
            db,
            category_cache,
            parent=level2,
            level=3,
            name_zh=row["基础动作_中文"],
            name_en=row["BaseMovement_EN"],
            sort_order=index,
        )

        exercise_key = (level3.id, row["动作名称_中文"], row["Movement_EN"])
        if exercise_key in exercise_keys:
            continue
        exercise_keys.add(exercise_key)

        db.add(
            Exercise(
                name=row["动作名称_中文"],
                alias=row["Movement_EN"] or None,
                base_category_id=level3.id,
                description=None,
                video_url=None,
                video_path=None,
                coaching_points=None,
                common_errors=None,
                notes=None,
                load_profile="general",
                default_increment=None,
                is_main_lift_candidate=False,
            )
        )

    db.commit()
    return preview


def get_category(db: Session, category_id: int) -> ExerciseCategory:
    category = db.query(ExerciseCategory).filter(ExerciseCategory.id == category_id).first()
    if not category:
        raise not_found("Exercise category not found")
    return category
