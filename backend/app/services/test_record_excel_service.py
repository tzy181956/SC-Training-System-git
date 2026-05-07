from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session, joinedload

from app.core.test_definition_defaults import DEFAULT_TEST_DEFINITION_CATALOG
from app.models import Athlete, TestRecord, TestTypeDefinition, User
from app.services import access_control_service, test_definition_service


ATHLETE_CODE_HEADER = "运动员编码"
ATHLETE_NAME_HEADER = "运动员姓名"
TEST_DATE_HEADER = "测试日期"
TEST_TYPE_HEADER = "测试类型"
METRIC_NAME_HEADER = "测试项目"
RESULT_VALUE_HEADER = "数值结果"
RESULT_TEXT_HEADER = "显示文本"
UNIT_HEADER = "单位"
NOTES_HEADER = "备注"

TEMPLATE_HEADERS = [
    ATHLETE_CODE_HEADER,
    ATHLETE_NAME_HEADER,
    TEST_DATE_HEADER,
    TEST_TYPE_HEADER,
    METRIC_NAME_HEADER,
    RESULT_VALUE_HEADER,
    RESULT_TEXT_HEADER,
    UNIT_HEADER,
    NOTES_HEADER,
]

REQUIRED_TEMPLATE_HEADERS = [
    ATHLETE_NAME_HEADER,
    TEST_DATE_HEADER,
    TEST_TYPE_HEADER,
    METRIC_NAME_HEADER,
    RESULT_VALUE_HEADER,
    UNIT_HEADER,
]

REQUIRED_HEADER_FONT = Font(bold=True, color="C62828")
OPTIONAL_HEADER_FONT = Font(bold=True)


@dataclass
class TestRecordImportSummary:
    total_rows: int
    imported_rows: int
    skipped_rows: int


def build_import_template_workbook(db: Session, user: User) -> bytes:
    visible_sport_id = access_control_service.resolve_visible_sport_id(user)
    athletes = _query_visible_athletes(db, visible_sport_id)
    definitions = test_definition_service.list_test_type_definitions(db, visible_sport_id=visible_sport_id)
    example_rows = _build_example_rows(definitions)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试数据导入模板"
    sheet.append(TEMPLATE_HEADERS)
    _style_template_header_row(sheet)

    example_athlete = athletes[0] if athletes else None
    example_name = example_athlete.full_name if example_athlete else "示例运动员"
    example_code = example_athlete.code if example_athlete and example_athlete.code else "ATH-000001"
    for row in example_rows:
        sheet.append(
            [
                example_code,
                example_name,
                row["test_date"],
                row["test_type"],
                row["metric_name"],
                row["result_value"],
                row["result_text"],
                row["unit"],
                row["notes"],
            ]
        )

    reference_sheet = workbook.create_sheet("运动员参考")
    reference_sheet.append([ATHLETE_CODE_HEADER, ATHLETE_NAME_HEADER, "队伍", "项目"])
    _style_bold_row(reference_sheet)
    for athlete in athletes:
        reference_sheet.append(
            [
                athlete.code,
                athlete.full_name,
                athlete.team.name if athlete.team else "",
                athlete.sport.name if athlete.sport else "",
            ]
        )

    definition_sheet = workbook.create_sheet("测试项目参考")
    definition_sheet.append([TEST_TYPE_HEADER, METRIC_NAME_HEADER, "默认单位", "范围", "项目", NOTES_HEADER])
    _style_bold_row(definition_sheet)
    for definition in definitions:
        scope_label = "系统" if definition.sport_id is None else "本项目"
        sport_name = definition.sport_name or ""
        if not definition.metrics:
            definition_sheet.append([definition.name, "", "", scope_label, sport_name, definition.notes or ""])
            continue
        for metric in definition.metrics:
            definition_sheet.append(
                [
                    definition.name,
                    metric.name,
                    metric.default_unit or "",
                    scope_label,
                    sport_name,
                    metric.notes or "",
                ]
            )

    instruction_sheet = workbook.create_sheet("填写说明")
    instruction_sheet.append(["字段", "说明"])
    _style_bold_row(instruction_sheet)
    instruction_sheet.append(["模板说明", "红色标题为必填项；黑色标题为选填项。"])
    instruction_sheet.append([ATHLETE_CODE_HEADER, "可留空；如系统中存在同名运动员，建议填写编码。"])
    instruction_sheet.append([ATHLETE_NAME_HEADER, "必填；必须与当前账号可管理范围内的运动员姓名一致。"])
    instruction_sheet.append([TEST_DATE_HEADER, "必填；建议填写 YYYY-MM-DD，例如 2026-03-14。"])
    instruction_sheet.append([TEST_TYPE_HEADER, "必填；必须使用“测试项目参考”工作表里的测试类型。"])
    instruction_sheet.append([METRIC_NAME_HEADER, "必填；必须使用“测试项目参考”工作表里的测试项目。"])
    instruction_sheet.append([RESULT_VALUE_HEADER, "必填；图表和统计使用该字段。"])
    instruction_sheet.append([RESULT_TEXT_HEADER, "选填；用于展示原始文本结果，例如 13'2''。"])
    instruction_sheet.append([UNIT_HEADER, "必填；例如 kg / cm / s / 次 / RSI。"])
    instruction_sheet.append([NOTES_HEADER, "选填。"])
    instruction_sheet.append(["权限提醒", "教练账号只能导入本项目运动员，并且只能使用系统项目和本项目私有项目。"])

    return _workbook_to_bytes(workbook)


def build_test_record_library_workbook(db: Session, user: User) -> bytes:
    visible_sport_id = access_control_service.resolve_visible_sport_id(user)
    query = (
        db.query(TestRecord)
        .options(
            joinedload(TestRecord.athlete).joinedload(Athlete.team),
            joinedload(TestRecord.athlete).joinedload(Athlete.sport),
        )
    )
    if visible_sport_id is not None:
        query = query.join(TestRecord.athlete).filter(Athlete.sport_id == visible_sport_id)
    records = query.order_by(TestRecord.test_date.desc(), TestRecord.id.desc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试数据总库"
    headers = [
        "记录ID",
        ATHLETE_CODE_HEADER,
        ATHLETE_NAME_HEADER,
        "项目",
        "队伍",
        TEST_DATE_HEADER,
        TEST_TYPE_HEADER,
        METRIC_NAME_HEADER,
        RESULT_VALUE_HEADER,
        RESULT_TEXT_HEADER,
        UNIT_HEADER,
        NOTES_HEADER,
        "创建时间",
    ]
    sheet.append(headers)
    _style_bold_row(sheet)

    for record in records:
        athlete = record.athlete
        sheet.append(
            [
                record.id,
                athlete.code if athlete else "",
                athlete.full_name if athlete else "",
                athlete.sport.name if athlete and athlete.sport else "",
                athlete.team.name if athlete and athlete.team else "",
                record.test_date.isoformat(),
                record.test_type,
                record.metric_name,
                record.result_value,
                record.result_text or "",
                record.unit,
                record.notes or "",
                record.created_at.isoformat() if record.created_at else "",
            ]
        )

    return _workbook_to_bytes(workbook)


def import_test_records_from_workbook(db: Session, file_bytes: bytes, user: User) -> TestRecordImportSummary:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    header_index = {header: index for index, header in enumerate(headers)}

    missing_headers = [header for header in REQUIRED_TEMPLATE_HEADERS if header not in header_index]
    if missing_headers:
        raise ValueError(f"导入模板缺少列：{', '.join(missing_headers)}")

    visible_sport_id = access_control_service.resolve_visible_sport_id(user)
    visible_athletes = _query_visible_athletes(db, visible_sport_id)
    all_athletes = db.query(Athlete).all()
    visible_athletes_by_code = {
        athlete.code.strip().upper(): athlete
        for athlete in visible_athletes
        if athlete.code and athlete.code.strip()
    }
    all_athletes_by_code = {
        athlete.code.strip().upper(): athlete
        for athlete in all_athletes
        if athlete.code and athlete.code.strip()
    }

    visible_athletes_by_name: dict[str, list[Athlete]] = {}
    all_athletes_by_name: dict[str, list[Athlete]] = {}
    for athlete in visible_athletes:
        visible_athletes_by_name.setdefault(athlete.full_name, []).append(athlete)
    for athlete in all_athletes:
        all_athletes_by_name.setdefault(athlete.full_name, []).append(athlete)

    visible_definitions = test_definition_service.list_test_type_definitions(db, visible_sport_id=visible_sport_id)
    all_definitions = test_definition_service.list_test_type_definitions(db, visible_sport_id=None)
    visible_pairs = {
        (definition.name, metric.name)
        for definition in visible_definitions
        for metric in definition.metrics
    }
    all_pairs = {
        (definition.name, metric.name)
        for definition in all_definitions
        for metric in definition.metrics
    }

    existing_query = db.query(TestRecord)
    if visible_sport_id is not None:
        existing_query = existing_query.join(TestRecord.athlete).filter(Athlete.sport_id == visible_sport_id)
    existing_keys = {
        (
            record.athlete_id,
            record.test_date.isoformat(),
            record.test_type,
            record.metric_name,
            float(record.result_value),
            record.result_text or "",
            record.unit,
        )
        for record in existing_query.all()
    }

    pending_records: list[TestRecord] = []
    errors: list[str] = []
    skipped_rows = 0
    total_rows = 0

    for row_number in range(2, sheet.max_row + 1):
        row_values = [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]
        if all(value in (None, "") for value in row_values):
            continue

        total_rows += 1
        try:
            athlete_name = _string_cell(row_values[header_index[ATHLETE_NAME_HEADER]], required=True)
            athlete_code = None
            if ATHLETE_CODE_HEADER in header_index:
                athlete_code = _optional_string_cell(row_values[header_index[ATHLETE_CODE_HEADER]])

            athlete = _resolve_athlete(
                athlete_code=athlete_code,
                athlete_name=athlete_name,
                visible_athletes_by_code=visible_athletes_by_code,
                visible_athletes_by_name=visible_athletes_by_name,
                all_athletes_by_code=all_athletes_by_code,
                all_athletes_by_name=all_athletes_by_name,
                visible_sport_id=visible_sport_id,
            )

            record_date = _parse_date(row_values[header_index[TEST_DATE_HEADER]])
            test_type = _string_cell(row_values[header_index[TEST_TYPE_HEADER]], required=True)
            metric_name = _string_cell(row_values[header_index[METRIC_NAME_HEADER]], required=True)
            _ensure_visible_definition_pair(
                test_type=test_type,
                metric_name=metric_name,
                visible_pairs=visible_pairs,
                all_pairs=all_pairs,
                visible_sport_id=visible_sport_id,
            )
            result_value = _float_cell(row_values[header_index[RESULT_VALUE_HEADER]])
            result_text = _optional_string_cell(row_values[header_index[RESULT_TEXT_HEADER]]) if RESULT_TEXT_HEADER in header_index else None
            unit = _string_cell(row_values[header_index[UNIT_HEADER]], required=True)
            notes = _optional_string_cell(row_values[header_index[NOTES_HEADER]]) if NOTES_HEADER in header_index else None

            duplicate_key = (
                athlete.id,
                record_date.isoformat(),
                test_type,
                metric_name,
                result_value,
                result_text or "",
                unit,
            )
            if duplicate_key in existing_keys:
                skipped_rows += 1
                continue

            pending_records.append(
                TestRecord(
                    athlete_id=athlete.id,
                    test_date=record_date,
                    test_type=test_type,
                    metric_name=metric_name,
                    result_value=result_value,
                    result_text=result_text,
                    unit=unit,
                    notes=notes,
                )
            )
            existing_keys.add(duplicate_key)
        except ValueError as exc:
            errors.append(f"第 {row_number} 行：{exc}")

    if errors:
        raise ValueError("；".join(errors[:10]))

    if pending_records:
        db.add_all(pending_records)
        db.commit()

    return TestRecordImportSummary(
        total_rows=total_rows,
        imported_rows=len(pending_records),
        skipped_rows=skipped_rows,
    )


def _style_template_header_row(sheet) -> None:
    required_headers = set(REQUIRED_TEMPLATE_HEADERS)
    for cell in sheet[1]:
        if cell.value in required_headers:
            cell.font = REQUIRED_HEADER_FONT
        else:
            cell.font = OPTIONAL_HEADER_FONT


def _style_bold_row(sheet) -> None:
    for cell in sheet[1]:
        cell.font = OPTIONAL_HEADER_FONT


def _resolve_athlete(
    *,
    athlete_code: str | None,
    athlete_name: str,
    visible_athletes_by_code: dict[str, Athlete],
    visible_athletes_by_name: dict[str, list[Athlete]],
    all_athletes_by_code: dict[str, Athlete],
    all_athletes_by_name: dict[str, list[Athlete]],
    visible_sport_id: int | None,
) -> Athlete:
    normalized_code = _normalize_optional_code(athlete_code)
    if normalized_code:
        athlete = visible_athletes_by_code.get(normalized_code)
        if athlete:
            if athlete.full_name != athlete_name:
                raise ValueError(
                    f"运动员编码 {normalized_code} 与姓名“{athlete_name}”不匹配，系统记录为“{athlete.full_name}”。"
                )
            return athlete
        if normalized_code in all_athletes_by_code and visible_sport_id is not None:
            raise ValueError(f"运动员编码 {normalized_code} 不在当前账号可管理项目内。")
        raise ValueError(f"系统中不存在运动员编码：{normalized_code}")

    visible_matches = visible_athletes_by_name.get(athlete_name, [])
    if visible_matches:
        if len(visible_matches) > 1:
            raise ValueError(f"运动员姓名“{athlete_name}”存在重名，请填写运动员编码后重试。")
        return visible_matches[0]

    if visible_sport_id is not None and all_athletes_by_name.get(athlete_name):
        raise ValueError(f"运动员“{athlete_name}”不在当前账号可管理项目内。")
    raise ValueError(f"系统中不存在运动员：{athlete_name}")


def _ensure_visible_definition_pair(
    *,
    test_type: str,
    metric_name: str,
    visible_pairs: set[tuple[str, str]],
    all_pairs: set[tuple[str, str]],
    visible_sport_id: int | None,
) -> None:
    pair = (test_type, metric_name)
    if pair in visible_pairs:
        return
    if visible_sport_id is not None and pair in all_pairs:
        raise ValueError(f"测试项目“{test_type} / {metric_name}”不在当前账号可用目录内。")
    raise ValueError(f"测试项目“{test_type} / {metric_name}”不存在，请先在测试项目目录中创建。")


def _query_visible_athletes(db: Session, visible_sport_id: int | None) -> list[Athlete]:
    query = (
        db.query(Athlete)
        .options(joinedload(Athlete.team), joinedload(Athlete.sport))
        .order_by(Athlete.full_name.asc(), Athlete.id.asc())
    )
    if visible_sport_id is not None:
        query = query.filter(Athlete.sport_id == visible_sport_id)
    return query.all()


def _normalize_optional_code(value: str | None) -> str | None:
    normalized = (value or "").strip().upper()
    return normalized or None


def _workbook_to_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _build_example_rows(definitions: list[TestTypeDefinition]) -> list[dict[str, str | float]]:
    rows: list[dict[str, str | float]] = []
    for definition in definitions:
        for metric in definition.metrics[:1]:
            rows.append(
                {
                    "test_date": "2026-03-14",
                    "test_type": definition.name,
                    "metric_name": metric.name,
                    "result_value": 55 if metric.default_unit == "kg" else 812.3 if metric.default_unit == "s" else 42,
                    "result_text": "13'2''" if metric.default_unit == "s" else "",
                    "unit": metric.default_unit or "",
                    "notes": "示例",
                }
            )
        if len(rows) >= 2:
            break

    if len(rows) >= 2:
        rows[1]["notes"] = "文本结果可选" if rows[1]["unit"] == "s" else "示例"
        return rows[:2]

    fallback_rows = []
    for definition in DEFAULT_TEST_DEFINITION_CATALOG:
        for metric in definition["metrics"][:1]:
            fallback_rows.append(
                {
                    "test_date": "2026-03-14",
                    "test_type": definition["name"],
                    "metric_name": metric["name"],
                    "result_value": 55 if metric["default_unit"] == "kg" else 812.3 if metric["default_unit"] == "s" else 42,
                    "result_text": "13'2''" if metric["default_unit"] == "s" else "",
                    "unit": metric["default_unit"] or "",
                    "notes": "示例",
                }
            )
        if len(fallback_rows) >= 2:
            break
    if len(fallback_rows) >= 2:
        fallback_rows[1]["notes"] = "文本结果可选" if fallback_rows[1]["unit"] == "s" else "示例"
    return fallback_rows[:2]


def _string_cell(value: object, *, required: bool = False) -> str:
    text = str(value).strip() if value is not None else ""
    if required and not text:
        raise ValueError("存在必填字段为空。")
    return text


def _optional_string_cell(value: object) -> str | None:
    text = _string_cell(value)
    return text or None


def _float_cell(value: object) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError("数值结果不能为空。")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"数值结果无法解析：{value}") from exc


def _parse_date(value: object) -> date:
    if value is None or str(value).strip() == "":
        raise ValueError("测试日期不能为空。")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip().replace("/", "-").replace(".", "-")
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"测试日期格式错误：{value}") from exc
