from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from re import sub

from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.models import Exercise, ExerciseCategory
from app.schemas.exercise_category import ExerciseImportPreview


EXOS_SOURCE_PATH = Path(r"C:\Users\tzy\Downloads\exos_action_library_tagged_for_codex.xlsx")

TAG_FIELD_MAP = {
    "标签_功能类型": "functionType",
    "标签_训练目标": "trainingGoal",
    "标签_动作区域": "bodyRegion",
    "标签_细分部位": "subBodyPart",
    "标签_主动作模式": "primaryPattern",
    "标签_动作模式补充": "secondaryPattern",
    "标签_方向属性": "direction",
    "标签_下肢主导": "lowerDominance",
    "标签_肢体组合": "limbCombination",
    "标签_侧别": "laterality",
    "标签_动力属性": "powerType",
    "标签_器械": "equipment",
    "标签_体位": "bodyPosition",
    "标签_应用场景": "usageScene",
    "标签_FMS项目": "fmsItem",
    "标签_FMS阶段": "fmsPhase",
    "标签_FMS等级": "fmsLevel",
}

ORIGINAL_ENGLISH_FIELDS = {
    "movementTypeCategoryEn": "一级分类_EN",
    "movementTypeEn": "二级分类_EN",
    "baseMovementEn": "基础动作_EN",
    "movementEn": "动作英文原名",
}


def _slugify(value: str) -> str:
    value = sub(r"[^a-zA-Z0-9\u4e00-\u9fff]+", "-", value.strip().lower())
    value = sub(r"-{2,}", "-", value).strip("-")
    return value or "item"


def _build_code(parent_code: str | None, name_zh: str, name_en: str | None) -> str:
    local = _slugify(name_en or name_zh)
    return f"{parent_code}/{local}" if parent_code else local


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_multi_values(value: object) -> list[str]:
    text = _normalize_text(value)
    if not text:
        return []
    parts = [_normalize_text(item) for item in text.split("|")]
    seen: set[str] = set()
    normalized: list[str] = []
    for item in parts:
        if not item or item in seen:
            continue
        seen.add(item)
        normalized.append(item)
    return normalized


def _load_sheet_rows(source_path: Path) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook["动作库_标签版"]
    headers = [_normalize_text(cell.value) for cell in next(worksheet.iter_rows(max_row=1))]
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {
            headers[index]: _normalize_text(value)
            for index, value in enumerate(values)
            if index < len(headers)
        }
        if row.get("动作名称"):
            rows.append(row)
    return headers, rows


def _normalize_row(row: dict[str, str]) -> dict:
    tags = {
        target_key: _split_multi_values(row.get(source_key))
        for source_key, target_key in TAG_FIELD_MAP.items()
    }
    search_keywords = _split_multi_values(row.get("标签_检索关键词"))
    for source_key in ("动作名称", "动作英文原名", "标签词条", "分类路径"):
        value = _normalize_text(row.get(source_key))
        if value and value not in search_keywords:
            search_keywords.append(value)

    return {
        "code": _normalize_text(row.get("动作ID建议")),
        "nameZh": _normalize_text(row.get("动作名称")),
        "nameEn": _normalize_text(row.get("动作英文原名")),
        "level1Category": _normalize_text(row.get("一级分类")),
        "level2Category": _normalize_text(row.get("二级分类")),
        "baseMovement": _normalize_text(row.get("基础动作")),
        "originalEnglishFields": {
            key: _normalize_text(row.get(source_key)) or None
            for key, source_key in ORIGINAL_ENGLISH_FIELDS.items()
        },
        "tags": tags,
        "searchKeywords": search_keywords,
        "tagText": _normalize_text(row.get("标签词条")),
        "categoryPath": _normalize_text(row.get("分类路径")),
        "rawRow": row,
    }


def _preview_rows(source_path: Path) -> tuple[list[dict], ExerciseImportPreview]:
    _, raw_rows = _load_sheet_rows(source_path)
    normalized_rows: list[dict] = []
    unique_codes: set[str] = set()
    level1 = set()
    level2 = set()
    level3 = set()
    skipped_duplicates = 0

    for row in raw_rows:
        normalized = _normalize_row(row)
        if not normalized["code"] or not normalized["nameZh"]:
            continue
        if normalized["code"] in unique_codes:
            skipped_duplicates += 1
            continue
        unique_codes.add(normalized["code"])
        normalized_rows.append(normalized)
        level1.add(normalized["level1Category"])
        level2.add((normalized["level1Category"], normalized["level2Category"]))
        level3.add((normalized["level1Category"], normalized["level2Category"], normalized["baseMovement"]))

    preview = ExerciseImportPreview(
        source_path=str(source_path),
        total_rows=len(raw_rows),
        valid_rows=len(normalized_rows),
        unique_codes=len(unique_codes),
        level1_categories=len(level1),
        level2_categories=len(level2),
        level3_categories=len(level3),
        exercises=len(normalized_rows),
        to_create=0,
        to_update=0,
        skipped_duplicates=skipped_duplicates,
    )
    return normalized_rows, preview


def list_categories(db: Session) -> list[ExerciseCategory]:
    return db.query(ExerciseCategory).order_by(ExerciseCategory.level, ExerciseCategory.sort_order, ExerciseCategory.name_zh).all()


def list_category_tree(db: Session) -> list[ExerciseCategory]:
    categories = list_categories(db)
    category_map = {category.id: category for category in categories}
    for category in categories:
        category.children = []
    roots: list[ExerciseCategory] = []
    for category in categories:
        if category.parent_id and category.parent_id in category_map:
            category_map[category.parent_id].children.append(category)
        else:
            roots.append(category)
    return roots


def _get_or_create_category(
    db: Session,
    cache: dict[tuple[int | None, int, str], ExerciseCategory],
    *,
    parent: ExerciseCategory | None,
    level: int,
    name_zh: str,
    name_en: str | None,
    sort_order: int,
) -> ExerciseCategory:
    cache_key = (parent.id if parent else None, level, name_zh)
    category = cache.get(cache_key)
    if category:
        return category

    category = (
        db.query(ExerciseCategory)
        .filter(
            ExerciseCategory.parent_id == (parent.id if parent else None),
            ExerciseCategory.level == level,
            ExerciseCategory.name_zh == name_zh,
        )
        .first()
    )
    if not category:
        category = ExerciseCategory(
            parent_id=parent.id if parent else None,
            level=level,
            name_zh=name_zh,
            name_en=name_en or None,
            code=_build_code(parent.code if parent else None, name_zh, name_en),
            sort_order=sort_order,
            is_system=True,
        )
        db.add(category)
        db.flush()

    cache[cache_key] = category
    return category


def _legacy_key(exercise: Exercise) -> tuple[str, str, int | None]:
    return (exercise.name or "", exercise.name_en or exercise.alias or "", exercise.base_category_id)


def _prepare_existing_maps(
    items: Iterable[Exercise],
) -> tuple[dict[str, Exercise], dict[tuple[str, str, int | None], list[Exercise]]]:
    by_code: dict[str, Exercise] = {}
    by_legacy: dict[tuple[str, str, int | None], list[Exercise]] = {}
    for item in items:
        if item.code:
            by_code[item.code] = item
        by_legacy.setdefault(_legacy_key(item), []).append(item)
    return by_code, by_legacy


def preview_exos_import(db: Session, source_path: Path | None = None) -> ExerciseImportPreview:
    path = source_path or EXOS_SOURCE_PATH
    normalized_rows, preview = _preview_rows(path)
    existing_items = db.query(Exercise).all()
    existing_by_code, existing_by_legacy = _prepare_existing_maps(existing_items)
    to_create = 0
    to_update = 0
    consumed_ids: set[int] = set()

    category_cache: dict[tuple[int | None, int, str], ExerciseCategory] = {}
    for index, row in enumerate(normalized_rows, start=1):
        level1 = _get_or_create_category(
            db,
            category_cache,
            parent=None,
            level=1,
            name_zh=row["level1Category"],
            name_en=row["originalEnglishFields"]["movementTypeCategoryEn"],
            sort_order=index,
        )
        level2 = _get_or_create_category(
            db,
            category_cache,
            parent=level1,
            level=2,
            name_zh=row["level2Category"],
            name_en=row["originalEnglishFields"]["movementTypeEn"],
            sort_order=index,
        )
        level3 = _get_or_create_category(
            db,
            category_cache,
            parent=level2,
            level=3,
            name_zh=row["baseMovement"],
            name_en=row["originalEnglishFields"]["baseMovementEn"],
            sort_order=index,
        )
        legacy_key = (row["nameZh"], row["nameEn"], level3.id)
        exercise = existing_by_code.get(row["code"])
        if exercise and exercise.id in consumed_ids:
            exercise = None
        if not exercise:
            candidates = existing_by_legacy.get(legacy_key, [])
            exercise = next((item for item in candidates if item.id not in consumed_ids), None)
        if exercise:
            consumed_ids.add(exercise.id)
            to_update += 1
        else:
            to_create += 1

    db.rollback()
    preview.to_create = to_create
    preview.to_update = to_update
    return preview


def _generate_manual_code(name: str, existing_codes: set[str]) -> str:
    base = f"CUSTOM_{_slugify(name).upper()}"
    code = base
    counter = 2
    while code in existing_codes:
        code = f"{base}_{counter}"
        counter += 1
    return code


def _cleanup_stale_imported_exercises(db: Session, active_codes: set[str]) -> None:
    stale_items = (
        db.query(Exercise)
        .filter(
            Exercise.source_type == "exos_excel",
            or_(Exercise.code.is_(None), ~Exercise.code.in_(active_codes)),
        )
        .all()
    )
    for item in stale_items:
        db.delete(item)
    db.flush()


def _cleanup_orphan_system_categories(db: Session) -> None:
    while True:
        orphan = (
            db.query(ExerciseCategory)
            .filter(ExerciseCategory.is_system.is_(True))
            .filter(~ExerciseCategory.exercises.any())
            .filter(~ExerciseCategory.children.any())
            .first()
        )
        if not orphan:
            break
        db.delete(orphan)
        db.flush()


def import_exos_library(db: Session, *, source_path: Path | None = None, replace_existing: bool = True) -> ExerciseImportPreview:
    path = source_path or EXOS_SOURCE_PATH
    normalized_rows, preview = _preview_rows(path)
    existing_items = db.query(Exercise).all()
    existing_by_code, existing_by_legacy = _prepare_existing_maps(existing_items)
    existing_codes = {item.code for item in existing_items if item.code}
    consumed_ids: set[int] = set()

    category_cache: dict[tuple[int | None, int, str], ExerciseCategory] = {}
    to_create = 0
    to_update = 0
    active_codes: set[str] = set()

    for index, row in enumerate(normalized_rows, start=1):
        level1 = _get_or_create_category(
            db,
            category_cache,
            parent=None,
            level=1,
            name_zh=row["level1Category"],
            name_en=row["originalEnglishFields"]["movementTypeCategoryEn"],
            sort_order=index,
        )
        level2 = _get_or_create_category(
            db,
            category_cache,
            parent=level1,
            level=2,
            name_zh=row["level2Category"],
            name_en=row["originalEnglishFields"]["movementTypeEn"],
            sort_order=index,
        )
        level3 = _get_or_create_category(
            db,
            category_cache,
            parent=level2,
            level=3,
            name_zh=row["baseMovement"],
            name_en=row["originalEnglishFields"]["baseMovementEn"],
            sort_order=index,
        )

        exercise = existing_by_code.get(row["code"])
        if exercise and exercise.id in consumed_ids:
            exercise = None
        if not exercise:
            candidates = existing_by_legacy.get((row["nameZh"], row["nameEn"], level3.id), [])
            exercise = next((item for item in candidates if item.id not in consumed_ids), None)

        if exercise:
            consumed_ids.add(exercise.id)
            to_update += 1
        else:
            to_create += 1
            exercise = Exercise(
                name=row["nameZh"],
                alias=row["nameEn"] or None,
                code=row["code"],
                source_type="exos_excel",
                is_main_lift_candidate=False,
            )
            db.add(exercise)

        exercise.name = row["nameZh"]
        exercise.alias = row["nameEn"] or None
        exercise.code = row["code"]
        exercise.source_type = "exos_excel"
        exercise.name_en = row["nameEn"] or None
        exercise.level1_category = row["level1Category"] or None
        exercise.level2_category = row["level2Category"] or None
        exercise.base_movement = row["baseMovement"] or None
        exercise.category_path = row["categoryPath"] or None
        exercise.original_english_fields = row["originalEnglishFields"]
        exercise.structured_tags = row["tags"]
        exercise.search_keywords = row["searchKeywords"]
        exercise.tag_text = row["tagText"] or None
        exercise.raw_row = row["rawRow"]
        exercise.base_category = level3
        existing_codes.add(row["code"])
        active_codes.add(row["code"])

    if replace_existing:
        _cleanup_stale_imported_exercises(db, active_codes)
        for item in db.query(Exercise).filter(or_(Exercise.code.is_(None), Exercise.code == "")).all():
            item.code = _generate_manual_code(item.name, existing_codes)
            existing_codes.add(item.code)
        _cleanup_orphan_system_categories(db)

    db.commit()
    preview.to_create = to_create
    preview.to_update = to_update
    return preview
