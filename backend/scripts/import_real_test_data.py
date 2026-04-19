from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))
VENV_SITE_PACKAGES = ROOT_DIR / ".venv" / "Lib" / "site-packages"
if VENV_SITE_PACKAGES.exists() and str(VENV_SITE_PACKAGES) not in sys.path:
    sys.path.append(str(VENV_SITE_PACKAGES))

from sqlalchemy import select

from app.core.schema_sync import ensure_runtime_schema
from app.core.database import SessionLocal
from app.models import (  # noqa: E402
    AssignmentItemOverride,
    Athlete,
    AthletePlanAssignment,
    SetRecord,
    Sport,
    Team,
    TestRecord,
    TrainingSession,
    TrainingSessionItem,
)


SOURCE_XLSX = Path(r"C:\Users\Tian Ziyu\OneDrive\上海\篮球中心\测试\20260314\测试结果带排名.xlsx")
TARGET_DATE = date(2026, 3, 14)
TARGET_SPORT_NAME = "篮球"
TARGET_SPORT_CODE = "basketball"
TARGET_TEAM_NAME = "女篮青年队"
TARGET_TEAM_CODE = "women-youth-basketball"

BODY_FIELDS = {
    "身高": ("height", "cm"),
    "体重": ("weight", "kg"),
    "体脂率": ("body_fat_percentage", "%"),
    "臂展": ("wingspan", "cm"),
    "站摸": ("standing_reach", "cm"),
}

METRIC_DEFINITIONS = {
    ("卧推", "重量"): ("力量测试", "卧推", "kg"),
    ("卧拉", "重量"): ("力量测试", "卧拉", "kg"),
    ("深蹲", "重量"): ("力量测试", "深蹲", "kg"),
    ("臀桥", "重量"): ("力量测试", "臀桥", "kg"),
    ("引体向上", "次数"): ("力量测试", "引体向上", "次"),
    ("反向跳", "高度"): ("体能测试", "反向跳", "cm"),
    ("静蹲跳", "高度"): ("体能测试", "静蹲跳", "cm"),
    ("直腿跳", "RSI"): ("体能测试", "直腿跳", "RSI"),
    ("助跑摸高", "高度"): ("体能测试", "助跑摸高", "cm"),
    ("原地摸高", "高度"): ("体能测试", "原地摸高", "cm"),
    ("30m跑", "10m"): ("速度测试", "30m跑-10m", "s"),
    ("30m跑", "30m"): ("速度测试", "30m跑-30m", "s"),
    ("505变向", "左腿"): ("速度测试", "505变向-左腿", "s"),
    ("505变向", "右腿"): ("速度测试", "505变向-右腿", "s"),
    ("505变向", "总用时"): ("速度测试", "505变向-总用时", "s"),
    ("限制区移动", "时间"): ("速度测试", "限制区移动", "s"),
    ("3000m", "秒数"): ("耐力测试", "3000m", "s"),
    ("17折*4", "秒数"): ("耐力测试", "17折*4", "s"),
}

TEXT_VALUE_COLUMNS = {
    ("3000m", "秒数"): ("3000m", "用时"),
    ("17折*4", "秒数"): ("17折*4", "平均用时"),
}
SKIP_NAME_MARKERS = ("平均", "汇总", "合计", "总计")

IGNORED_SUB_HEADERS = {"排名"}
IGNORED_TOP_HEADERS = {"总分", "总排名"}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_header_rows(ws) -> tuple[dict[int, tuple[str, str]], dict[tuple[str, str], int]]:
    merged_lookup: dict[int, str] = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        if min_row == 1:
            header_value = normalize_text(ws.cell(min_row, min_col).value)
            for column in range(min_col, max_col + 1):
                merged_lookup[column] = header_value

    columns: dict[int, tuple[str, str]] = {}
    reverse_lookup: dict[tuple[str, str], int] = {}
    for column in range(1, ws.max_column + 1):
        top = normalize_text(ws.cell(1, column).value) or merged_lookup.get(column, "")
        sub = normalize_text(ws.cell(2, column).value)
        if top in IGNORED_TOP_HEADERS or sub in IGNORED_SUB_HEADERS:
            continue
        columns[column] = (top, sub)
        reverse_lookup[(top, sub)] = column
    return columns, reverse_lookup


def parse_time_text(value: str) -> float:
    text = normalize_text(value)
    if not text:
        raise ValueError("empty time text")

    minutes = 0
    if "′" in text:
        minute_text, text = text.split("′", 1)
        minutes = int(minute_text)

    if "″" in text:
        second_text, fraction_text = text.split("″", 1)
        seconds = int(second_text or 0)
        fraction = 0.0
        if fraction_text:
            fraction = float(f"0.{fraction_text}")
        return round(minutes * 60 + seconds + fraction, 2)

    return round(minutes * 60 + float(text), 2)


def ensure_target_org(db) -> tuple[Sport, Team]:
    sport = db.execute(select(Sport).where(Sport.name == TARGET_SPORT_NAME)).scalar_one_or_none()
    if not sport:
        sport = Sport(name=TARGET_SPORT_NAME, code=TARGET_SPORT_CODE, notes="真实导入数据所属项目")
        db.add(sport)
        db.flush()

    team = db.execute(select(Team).where(Team.sport_id == sport.id, Team.name == TARGET_TEAM_NAME)).scalar_one_or_none()
    if not team:
        team = Team(sport_id=sport.id, name=TARGET_TEAM_NAME, code=TARGET_TEAM_CODE, notes="真实导入女篮青年队")
        db.add(team)
        db.flush()

    return sport, team


def clear_business_data(db) -> None:
    db.query(SetRecord).delete()
    db.query(TrainingSessionItem).delete()
    db.query(TrainingSession).delete()
    db.query(AssignmentItemOverride).delete()
    db.query(AthletePlanAssignment).delete()
    db.query(TestRecord).delete()
    db.query(Athlete).delete()
    db.commit()


def build_or_update_athlete(db, team: Team, row_map: dict[str, object]) -> Athlete:
    full_name = normalize_text(row_map["姓名"])
    athlete = db.execute(select(Athlete).where(Athlete.full_name == full_name)).scalar_one_or_none()
    payload = {
        "user_id": None,
        "sport_id": team.sport_id,
        "team_id": team.id,
        "full_name": full_name,
        "position": None,
        "training_level": None,
        "height": row_map.get("身高"),
        "weight": row_map.get("体重"),
        "body_fat_percentage": row_map.get("体脂率"),
        "wingspan": row_map.get("臂展"),
        "standing_reach": row_map.get("站摸"),
        "notes": None,
        "is_active": True,
    }
    if athlete:
        for key, value in payload.items():
            setattr(athlete, key, value)
    else:
        athlete = Athlete(**payload)
        db.add(athlete)
        db.flush()
    return athlete


def add_test_record(db, athlete: Athlete, test_type: str, metric_name: str, unit: str, result_value: float, result_text: str | None = None) -> None:
    db.add(
        TestRecord(
            athlete_id=athlete.id,
            test_date=TARGET_DATE,
            test_type=test_type,
            metric_name=metric_name,
            result_value=result_value,
            result_text=result_text,
            unit=unit,
            notes=None,
        )
    )


def import_workbook(db, workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    columns, reverse_lookup = parse_header_rows(worksheet)

    sport, team = ensure_target_org(db)
    clear_business_data(db)
    sport, team = ensure_target_org(db)

    seen_names: set[str] = set()
    for row_index in range(3, worksheet.max_row + 1):
        row_values = {columns[column][0] or get_column_letter(column): worksheet.cell(row_index, column).value for column in columns}
        full_name = normalize_text(row_values.get("姓名"))
        if not full_name:
            continue
        if any(marker in full_name for marker in SKIP_NAME_MARKERS):
            continue
        if full_name in seen_names:
            raise ValueError(f"Duplicate athlete name in Excel: {full_name}")
        seen_names.add(full_name)

        athlete = build_or_update_athlete(db, team, row_values)

        for body_metric, (_, unit) in BODY_FIELDS.items():
            value = row_values.get(body_metric)
            if value is None:
                continue
            add_test_record(db, athlete, "基础身体", body_metric, unit, float(value))

        for header_key, (test_type, metric_name, unit) in METRIC_DEFINITIONS.items():
            column = reverse_lookup.get(header_key)
            if not column:
                continue
            raw_value = worksheet.cell(row_index, column).value
            if raw_value is None or normalize_text(raw_value) == "":
                continue

            result_value = float(raw_value)
            result_text = None
            if header_key in TEXT_VALUE_COLUMNS:
                text_column = reverse_lookup.get(TEXT_VALUE_COLUMNS[header_key])
                text_value = worksheet.cell(row_index, text_column).value if text_column else None
                result_text = normalize_text(text_value) or None
                if result_text:
                    parsed_seconds = parse_time_text(result_text)
                    if abs(parsed_seconds - result_value) > 0.11:
                        raise ValueError(
                            f"Time parse mismatch for {full_name} / {metric_name}: text={result_text}, seconds={result_value}"
                        )

            add_test_record(db, athlete, test_type, metric_name, unit, result_value, result_text)

    db.commit()


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Workbook not found: {SOURCE_XLSX}")

    ensure_runtime_schema()
    with SessionLocal() as db:
        import_workbook(db, SOURCE_XLSX)
    print("Real athlete and test data imported.")


if __name__ == "__main__":
    main()
